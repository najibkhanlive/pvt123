"""
Excel Parser: Reads ITSM tickets from uploaded Excel files.
Parses the 'Incidents' sheet and maps columns to ticket dictionaries.
"""

import openpyxl
from typing import List, Dict, Optional
from io import BytesIO


# Expected column mappings (Excel header → internal key)
COLUMN_MAP = {
    "Number": "Number",
    "Opened": "Opened",
    "Short description": "Short description",
    "Type": "Type",
    "Affected User": "Affected User",
    "Opened by": "Opened by",
    "Priority": "Priority",
    "State": "State",
    "Categorization": "Categorization",
    "Assignment group": "Assignment group",
    "Assigned to": "Assigned to",
    "Updated": "Updated",
    "Updated by": "Updated by",
    "Initial 1st level": "Initial 1st level",
    "2nd lvl Assignment Group": "2nd lvl Assignment Group",
    "Subcategory": "Subcategory",
    "Reassignment count": "Reassignment count",
    "Reassignment Reason": "Reassignment Reason",
    "Recent Assignment Group": "Recent Assignment Group",
    "Work notes": "Work notes",
    "Resolution notes": "Resolution notes",
    "L2 Assignment Group count": "L2 Assignment Group count",
    "Short description (automatically translated)": "Short description (translated)",
    "Country": "Country",
    "Resolved by": "Resolved by",
}


def parse_excel(file_bytes: bytes, sheet_name: str = "Incidents") -> List[Dict]:
    """
    Parse an Excel file and return a list of ticket dictionaries.
    
    Args:
        file_bytes: Raw bytes of the Excel file
        sheet_name: Name of the sheet to read (default: 'Incidents')
    
    Returns:
        List of ticket dictionaries with all available fields
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    
    # Try to find the sheet (case-insensitive)
    target_sheet = None
    for name in wb.sheetnames:
        if name.lower() == sheet_name.lower():
            target_sheet = name
            break
    
    if target_sheet is None:
        # Fallback to first sheet
        target_sheet = wb.sheetnames[0]
    
    ws = wb[target_sheet]
    
    # Read headers from first row
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    headers = [str(h).strip() if h else "" for h in rows[0]]
    
    # Map header indices
    header_indices = {}
    for idx, header in enumerate(headers):
        # Try exact match first
        if header in COLUMN_MAP:
            header_indices[COLUMN_MAP[header]] = idx
        else:
            # Try fuzzy match (lowercase, stripped)
            header_lower = header.lower().strip()
            for excel_key, internal_key in COLUMN_MAP.items():
                if excel_key.lower().strip() == header_lower:
                    header_indices[internal_key] = idx
                    break
    
    # Parse data rows
    tickets = []
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        
        ticket = {}
        for internal_key, col_idx in header_indices.items():
            if col_idx < len(row):
                value = row[col_idx]
                ticket[internal_key] = str(value).strip() if value is not None else ""
            else:
                ticket[internal_key] = ""
        
        # Skip rows without a ticket number
        if not ticket.get("Number", "").strip():
            continue
        
        tickets.append(ticket)
    
    wb.close()
    return tickets


def validate_tickets(tickets: List[Dict]) -> Dict:
    """
    Validate parsed tickets and return a summary.
    """
    if not tickets:
        return {
            "valid": False,
            "error": "No tickets found in the Excel file",
            "count": 0,
        }
    
    # Check for required fields
    required = ["Number", "Work notes"]
    missing_fields = []
    for field in required:
        if not any(t.get(field, "").strip() for t in tickets):
            missing_fields.append(field)
    
    warnings = []
    if missing_fields:
        warnings.append(f"Missing key fields: {', '.join(missing_fields)}")
    
    # Count tickets with work notes
    with_notes = sum(1 for t in tickets if t.get("Work notes", "").strip())
    
    return {
        "valid": True,
        "count": len(tickets),
        "with_work_notes": with_notes,
        "fields_found": list(tickets[0].keys()) if tickets else [],
        "warnings": warnings,
        "ticket_ids": [t.get("Number", "Unknown") for t in tickets],
    }


def get_ticket_summary(ticket: Dict) -> Dict:
    """Return a lightweight summary of a ticket for the frontend listing."""
    return {
        "Number": ticket.get("Number", ""),
        "Short description": ticket.get("Short description", "")[:100],
        "Priority": ticket.get("Priority", ""),
        "State": ticket.get("State", ""),
        "Categorization": ticket.get("Categorization", ""),
        "Affected User": ticket.get("Affected User", ""),
        "Country": ticket.get("Country", ""),
        "Opened": ticket.get("Opened", ""),
    }